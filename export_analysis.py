import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.styles.fonts import Font
from openpyxl.worksheet.hyperlink import Hyperlink

# Get current date and time for filename
current_datetime = datetime.datetime.now()
date_str = current_datetime.strftime("%Y-%m-%d")
time_str = current_datetime.strftime("%H%M")
excel_filename = f"{date_str}-{time_str}-stocks.xlsx"

# Analysis results from our previous runs
stock_analysis = {
    "MSFT": {
        "agent": "Benjamin Graham",
        "signal": "SHORT",
        "current_price": 385.24,
        "target_price": 308.19,  # 20% downside
        "confidence": 75.0,
        "reasoning": "The stock trades at $337.97 per share, which is 1,764% above the Graham Number of $18.12 and 9,526% above NCAV per share of $3.51, completely eliminating the margin of safety. While the current ratio of 2.47 exceeds Graham's 2.0 threshold and debt ratio of 0.54 remains within acceptable bounds, the extreme valuation disconnect cannot be ignored."
    },
    "NVDA": {
        "agent": "Benjamin Graham",
        "signal": "NEUTRAL",
        "current_price": 106.84,
        "target_price": None,
        "confidence": 35.0,
        "reasoning": "While NVIDIA displays strong financial health (current ratio 2.57, conservative debt ratio 0.38), its extreme valuation is concerning. However, as the leader in AI chips, it has significant upside potential that offsets the valuation concerns."
    },
    "NEM": {
        "agent": "Benjamin Graham",
        "signal": "NEUTRAL (WATCH)",
        "current_price": 54.51,
        "target_price": None,
        "confidence": 35.0,
        "reasoning": "NEM has excellent financial strength (current ratio 3.52, debt ratio 0.43) but mixed earnings history. As a gold mining company, it may benefit from gold's rising price trends. Worth monitoring closely as global trade tensions might further boost gold prices."
    },
    "JPM": {
        "agent": "Benjamin Graham",
        "signal": "BEARISH (No Position)",
        "current_price": 242.88,
        "target_price": None,
        "confidence": 75.0,
        "reasoning": "Despite JPMorgan's strong Q1 profit, the high debt ratio (0.89) and premium to Graham Number (45.8%) make it unattractive from a value perspective. However, banks with strong trading desks like JPM might benefit from market volatility."
    },
    "META": {
        "agent": "Benjamin Graham",
        "signal": "NEUTRAL",
        "current_price": 549.02,
        "target_price": None,
        "confidence": 35.0,
        "reasoning": "META exhibits strong financial strength with a current ratio of 11.25 and a conservative debt ratio of 0.11. However, the stock trades at a significant premium to intrinsic value metrics, offering no margin of safety."
    },
    "CVX": {
        "agent": "Benjamin Graham",
        "signal": "NEUTRAL",
        "current_price": 138.12,
        "target_price": None,
        "confidence": 45.0,
        "reasoning": "CVX fails to meet Graham's margin of safety requirements. While the debt ratio of 0.42 remains below Graham's 0.50 threshold and dividend history provides some stability, the weak current ratio of 1.35 falls short of Graham's 2.0 minimum for liquidity."
    },
    "AAPL": {
        "agent": "Benjamin Graham",
        "signal": "NEUTRAL",
        "current_price": 206.32,
        "target_price": None,
        "confidence": 45.0,
        "reasoning": "The stock trades at an 87% premium to the Graham Number, eliminating the margin of safety. Liquidity metrics (current ratio = 1.11) fall short of Graham's 2.0 threshold, and the debt ratio of 0.59 suggests elevated leverage for a defensive investment."
    },
    "COST": {
        "agent": "Benjamin Graham",
        "signal": "NEUTRAL",
        "current_price": 972.0,
        "target_price": None,
        "confidence": 40.0,
        "reasoning": "COST fails to meet key Graham criteria for a margin of safety or financial strength. The stock trades at a premium to its Graham Number, offering no valuation cushion. Financial metrics show weakness with a current ratio of 1.01 and a debt ratio of 0.67."
    },
    "BA": {
        "agent": "Benjamin Graham",
        "signal": "BEARISH",
        "current_price": 175.41,
        "target_price": None,
        "confidence": 80.0,
        "reasoning": "The stock exhibits multiple fundamental weaknesses. Financial strength is precarious: the current ratio of 1.35 falls below Graham's 2.0 liquidity requirement, while the debt ratio of 0.93 indicates excessive leverage. Earnings instability is evident with negative EPS in multiple periods."
    }
}

# Portfolio recommendations
portfolio_recommendations = {
    "primary_position": {
        "ticker": "MSFT", 
        "action": "SHORT",
        "quantity": 51,
        "entry_price": 385.24,
        "target_exit": 308.19,
        "stop_loss": 424.0
    },
    "cash_position": "80% of portfolio to preserve capital during market volatility",
    "watchlist": ["NEM", "JPM", "CVX"],
    "risk_management": [
        "Set stop-loss on MSFT short at $424 (10% above entry)",
        "Implement trailing stops to lock in profits",
        "Maintain position sizes below 20% of portfolio value",
        "Monitor trade war developments closely"
    ]
}

# Economic factors to monitor with news sources, dates, summaries and URLs
economic_news = [
    {
        "factor": "Trade War Developments",
        "date": "April 5, 2025",
        "source": "Currency News, Yahoo Finance",
        "summary": "Historic $2 Trillion was wiped off the US stock market as trade tensions escalate. The S&P 500 posted a 4.8% decline, sharpest slide since 2020.",
        "detail": "Watch for escalation or de-escalation of tariffs, especially between US-China. The US has implemented tariffs of 10% for most trading partners and 145% for China.",
        "ref_url_1": "https://www.currencynews.co.uk/forecast/20250405-42696_pound-dives-vs-euro-and-dollar-as-2-trillion-wiped-off-us-stock-market.html",
        "ref_url_2": "https://finance.yahoo.com/news/stocks-bonds-other-markets-fared-190055002.html",
        "ref_url_3": None
    },
    {
        "factor": "Federal Reserve Policy",
        "date": "April 2024",
        "source": "IMF World Economic Outlook",
        "summary": "Global inflation is forecast to decline steadily, from 6.8% in 2023 to 5.9% in 2024 and 4.5% in 2025, with advanced economies returning to inflation targets sooner.",
        "detail": "Monitor for interest rate decisions; high rates for longer than expected would pressure tech valuations further.",
        "ref_url_1": "https://www.imf.org/en/Publications/WEO/Issues/2024/04/16/world-economic-outlook-april-2024",
        "ref_url_2": None,
        "ref_url_3": None
    },
    {
        "factor": "Gold Prices",
        "date": "April 2025",
        "source": "Yahoo Finance, JPMorgan",
        "summary": "Gold reached record highs near $3,343 per Troy ounce. JP Morgan analysts see gold prices crossing $4,000/oz by Q2 2026.",
        "detail": "Track gold performance as a barometer of global uncertainty.",
        "ref_url_1": "https://finance.yahoo.com/news/stocks-bonds-other-markets-fared-190055002.html",
        "ref_url_2": "https://www.marketscreener.com/quote/stock/JPMORGAN-CHASE-CO-37468997/news/JP-Morgan-see-gold-prices-crossing-4-000-oz-by-Q2-2026-49680068/",
        "ref_url_3": None
    },
    {
        "factor": "Earnings Reports",
        "date": "April 11, 2025",
        "source": "Market Screener",
        "summary": "JPMorgan logged Q1 profit of $14.6 billion as CEO warns of uncertainty over global trade. Trading desk thrived in the volatile market.",
        "detail": "Pay special attention to how companies are being affected by trade tensions in their quarterly reports.",
        "ref_url_1": "https://www.marketscreener.com/quote/stock/JPMORGAN-CHASE-CO-37468997/news/JPMorgan-logs-Q1-profit-of-14-6-billion-as-CEO-warns-of-uncertainty-over-global-trade-other-events-49597780/",
        "ref_url_2": None,
        "ref_url_3": None
    },
    {
        "factor": "Market Volatility",
        "date": "April 2024",
        "source": "JMG Financial Group, Schroders",
        "summary": "Stocks ended April lower as benchmark indices endured their first downturn in several months due to Federal Reserve's intent to hold interest rates at a two-decade high.",
        "detail": "Consumer confidence fell in April to its lowest level since 2022. Labor costs increased the most in a year, driven higher by wage pressures that are helping to push inflation higher.",
        "ref_url_1": "https://www.jmgfinancial.com/why-did-stocks-pull-back-in-april/",
        "ref_url_2": "https://www.schroders.com/en/global/individual/insights/monthly-markets-review---april-2024/",
        "ref_url_3": None
    }
]

# Create DataFrames
df_analysis = pd.DataFrame.from_dict(stock_analysis, orient='index')
df_analysis.reset_index(inplace=True)
df_analysis.rename(columns={'index': 'Ticker'}, inplace=True)

# Create a DataFrame for the portfolio recommendation
df_portfolio = pd.DataFrame([portfolio_recommendations['primary_position']])

# Create a DataFrame for the watchlist
df_watchlist = pd.DataFrame({"Watchlist Tickers": portfolio_recommendations['watchlist']})

# Create a DataFrame for risk management
df_risk = pd.DataFrame({"Risk Management": portfolio_recommendations['risk_management']})

# Create a DataFrame for economic factors with news sources
df_economic = pd.DataFrame(economic_news)

# Export to Excel with the date-based filename
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df_analysis.to_excel(writer, sheet_name='Stock Analysis', index=False)
    df_portfolio.to_excel(writer, sheet_name='Portfolio Recommendation', index=False)
    df_watchlist.to_excel(writer, sheet_name='Watchlist', index=False)
    df_risk.to_excel(writer, sheet_name='Risk Management', index=False)
    df_economic.to_excel(writer, sheet_name='Economic Factors', index=False)
    
    # Apply formatting after writing data
    workbook = writer.book
    
    # Format Stock Analysis sheet
    ws_analysis = writer.sheets['Stock Analysis']
    for col in ['B', 'F']:  # Columns with potentially long text
        for row in range(2, len(df_analysis) + 2):
            cell = f"{col}{row}"
            ws_analysis[cell].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws_analysis.column_dimensions['A'].width = 10  # Ticker
    ws_analysis.column_dimensions['B'].width = 15  # Agent
    ws_analysis.column_dimensions['C'].width = 15  # Signal
    ws_analysis.column_dimensions['D'].width = 12  # Current price
    ws_analysis.column_dimensions['E'].width = 12  # Target price
    ws_analysis.column_dimensions['F'].width = 12  # Confidence
    ws_analysis.column_dimensions['G'].width = 60  # Reasoning
    
    # Add header formatting
    header_font = Font(bold=True)
    header_border = Border(bottom=Side(style='thin'))
    for col in range(1, df_analysis.shape[1] + 1):
        cell = ws_analysis.cell(row=1, column=col)
        cell.font = header_font
        cell.border = header_border
    
    # Format Economic Factors sheet
    ws_economic = writer.sheets['Economic Factors']
    for col in ['B', 'D', 'E']:  # Columns with potentially long text
        for row in range(2, len(df_economic) + 2):
            cell = f"{col}{row}"
            ws_economic[cell].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws_economic.column_dimensions['A'].width = 20  # Factor
    ws_economic.column_dimensions['B'].width = 15  # Date
    ws_economic.column_dimensions['C'].width = 20  # Source
    ws_economic.column_dimensions['D'].width = 30  # Summary
    ws_economic.column_dimensions['E'].width = 40  # Detail
    ws_economic.column_dimensions['F'].width = 30  # Reference URL 1
    ws_economic.column_dimensions['G'].width = 30  # Reference URL 2
    ws_economic.column_dimensions['H'].width = 30  # Reference URL 3
    
    # Add header formatting to economic factors
    for col in range(1, df_economic.shape[1] + 1):
        cell = ws_economic.cell(row=1, column=col)
        cell.font = header_font
        cell.border = header_border
    
    # Make URLs clickable
    hyperlink_font = Font(color="0000FF", underline="single")
    for row in range(2, len(df_economic) + 2):
        for col_idx, col_letter in enumerate(['F', 'G', 'H']):  # URL columns
            cell = f"{col_letter}{row}"
            url = ws_economic[cell].value
            if url and isinstance(url, str):
                # Create hyperlink
                cell_obj = ws_economic[cell]
                cell_obj.hyperlink = url
                cell_obj.font = hyperlink_font
                cell_obj.value = f"Reference {col_idx+1}"
    
    # Format row heights for wrapped text
    for sheet in [ws_analysis, ws_economic]:
        for row in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 70

print(f"Analysis exported to {excel_filename}") 